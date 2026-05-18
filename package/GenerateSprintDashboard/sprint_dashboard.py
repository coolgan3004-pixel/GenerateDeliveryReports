"""
Sprint Dashboard Generator & Email Sender
==========================================
Reads config from config.json (same folder as this script).

Discovery:
  1. Recursively scans workbooks_dir for files ending with 'Sprint Metrics.xlsx'
     (also matches 'Sprint_Metrics.xlsx' — space or underscore both work)
  2. Other xlsx files in the same folders are ignored
  3. Folders without a matching file are silently ignored (project ended)
  4. Errors raised only when a file IS found but has no valid sprint
  5. Application name = filename minus 'Sprint Metrics.xlsx', trailing
     hyphens/spaces stripped  (e.g. 'BOH - Sprint Metrics.xlsx' → 'BOH')

Column layout: G=StoryID  H=Description  I=Status (fixed, 0-indexed col 8)

Requirements:  pip install openpyxl pywin32
Usage:
    python sprint_dashboard.py                     # Generate only
    python sprint_dashboard.py --send              # Generate + send via Outlook
    python sprint_dashboard.py --config other.json # Alternate config
"""

import os, sys, json, argparse, re
import html as html_mod
from datetime import datetime, timedelta
from pathlib import Path


def load_config(path):
    if not os.path.exists(path):
        print(f"ERROR: {path} not found"); sys.exit(1)
    with open(path, "r", encoding="utf-8") as f:
        c = json.load(f)
    for k in ["workbooks_dir", "output_dir"]:
        if not c.get(k): print(f"ERROR: '{k}' required in config.json"); sys.exit(1)
    c.setdefault("template_file", "dashboard_template.html")
    c.setdefault("email", {})
    c["email"].setdefault("to", ""); c["email"].setdefault("cc", "")
    c["email"].setdefault("subject", "Sprint Dashboard — {date}")
    c["email"].setdefault("attach_interactive", True)
    c.setdefault("status_map", {}); c.setdefault("trivial_review_values", [])
    return c


def norm_status(raw, smap):
    return smap.get(str(raw).strip().lower(), "Other")

def safe_float(v):
    try:
        f = float(v)
        return f if f == f else 0
    except (TypeError, ValueError):
        return 0

def display_name_from_file(filepath):
    """Derive application display name from filename.
    'MSB Accounting - Sprint Metrics.xlsx' → 'MSB Accounting'
    'BOH__Sprint_Metrics.xlsx'             → 'BOH'
    'FOH Wave - Sprint Metrics.xlsx'       → 'FOH Wave'
    """
    name = Path(filepath).name
    # Remove 'Sprint Metrics.xlsx' or 'Sprint_Metrics.xlsx' (with optional space/underscore)
    name = re.sub(r'Sprint[\s_]*Metrics\.xlsx$', '', name, flags=re.IGNORECASE)
    # Strip trailing underscores, hyphens, spaces, dots
    name = name.rstrip(' _-.')
    # Replace underscores with spaces
    name = name.replace('_', ' ')
    # Collapse multiple spaces
    name = re.sub(r'\s+', ' ', name).strip()
    return name or Path(filepath).stem


# Fixed column indices (0-based)
COL_DEVELOPER = 5   # F
COL_STORY_ID  = 6   # G
COL_DESC      = 7   # H
COL_STATUS    = 8   # I


def extract_team(filepath, today, smap, trivials):
    """Returns (team_dict | None, error_string | None)."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheets = wb.sheetnames

    dash = None
    for s in sheets:
        if s.lower() == "dashboard": dash = s; break
    if not dash:
        for s in sheets:
            if "dashboard" in s.lower():
                ws = wb[s]
                r1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for i, v in enumerate(r1):
                    if isinstance(v, str) and v.strip().lower() == "current sprint" and i+1 < len(r1):
                        if r1[i+1] is not None and str(r1[i+1]).strip() not in ("", "nan"):
                            dash = s; break
                if dash: break
    if not dash:
        wb.close(); return None, "No active dashboard sheet found"

    ws = wb[dash]
    r1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    info = {}
    for i, v in enumerate(r1):
        if not isinstance(v, str): continue
        vl = v.strip().lower()
        if vl == "current sprint" and i+1 < len(r1) and r1[i+1] is not None:
            info["sprint"] = str(r1[i+1]).strip()
        elif vl == "current sprint sheet" and i+1 < len(r1) and r1[i+1] is not None:
            info["sheet"] = str(r1[i+1]).strip()
        elif vl == "sprint start date" and i+1 < len(r1):
            info["start"] = r1[i+1]
        elif vl == "sprint end date" and i+1 < len(r1):
            info["end"] = r1[i+1]
        elif "application" in vl and "name" in vl and i+1 < len(r1):
            info["app"] = str(r1[i+1]).strip()

    if not info.get("sprint") or info["sprint"] in ("nan", "None", ""):
        wb.close(); return None, "Current Sprint is empty on dashboard"

    start_dt, end_dt = info.get("start"), info.get("end")
    if isinstance(start_dt, (int, float)):
        start_dt = datetime(1899, 12, 30) + timedelta(days=int(start_dt))
    if isinstance(end_dt, (int, float)):
        end_dt = datetime(1899, 12, 30) + timedelta(days=int(end_dt))
    if not isinstance(end_dt, datetime):
        wb.close()
        return None, f"Sprint '{info['sprint']}' — invalid end date: {info.get('end')}"

    ss = info.get("sheet", "").strip()
    if not ss:
        wb.close(); return None, "Dashboard has no 'Current Sprint Sheet' value"

    # Build trimmed→actual sheet name lookup to handle trailing spaces
    sheet_lookup = {s.strip(): s for s in sheets}
    if ss in sheet_lookup:
        ss_actual = sheet_lookup[ss]
    else:
        available = [s.strip() for s in sheets if "-Status" in s]
        wb.close()
        return None, f"Sheet '{ss}' not found — typo? (available: {', '.join(available) if available else 'none'})"

    # Stories via fixed columns
    stories = []
    ws_s = wb[ss_actual]
    rows = list(ws_s.iter_rows(values_only=False))
    if rows:
        headers = [str(c.value).strip() if c.value else "" for c in rows[0]]
        headers_lower = [h.lower() for h in headers]
        est_idx = act_idx = None
        review_idxs, pov_idxs = [], []
        for i, hl in enumerate(headers_lower):
            if "estimated" in hl: est_idx = i
            elif hl.startswith("actual") and ("sp" in hl or "hrs" in hl): act_idx = i
            elif "code review comment" in hl or "peer tested defects" in hl or "peer tested comments" in hl:
                review_idxs.append(i)
            elif "pov" in hl: pov_idxs.append(i)

        for row in rows[1:]:
            v = [c.value for c in row]
            if len(v) <= COL_STATUS: continue
            sid = str(v[COL_STORY_ID]).strip() if v[COL_STORY_ID] is not None else ""
            if not sid or sid in ("nan", "None"): continue
            raw_st = str(v[COL_STATUS]).strip() if v[COL_STATUS] is not None else ""
            s = {
                "id": sid,
                "desc": str(v[COL_DESC])[:120] if v[COL_DESC] is not None else "",
                "status": norm_status(raw_st, smap),
                "est_sp": safe_float(v[est_idx]) if est_idx and est_idx < len(v) else 0,
                "act_sp": safe_float(v[act_idx]) if act_idx and act_idx < len(v) else 0,
                "dev": str(v[COL_DEVELOPER])[:40].strip() if v[COL_DEVELOPER] is not None else "",
                "review_flags": [],
            }
            for ri in review_idxs + pov_idxs:
                if ri < len(v) and v[ri] is not None:
                    rv = str(v[ri]).strip()
                    if rv.lower() not in trivials:
                        s["review_flags"].append(("[POV] " if ri in pov_idxs else "") + rv[:200])
            stories.append(s)

    # Defects
    ds = ss.replace("-Status", "-Defect").strip()
    ds_actual = sheet_lookup.get(ds)
    defects = []
    if ds_actual:
        ws_d = wb[ds_actual]
        dr = list(ws_d.iter_rows(values_only=False))
        if dr:
            dh = [str(c.value).strip().lower() if c.value else "" for c in dr[0]]
            for row in dr[1:]:
                v = [c.value for c in row]
                d = {}
                for i, h in enumerate(dh):
                    if i >= len(v) or v[i] is None: continue
                    vs = str(v[i]).strip()
                    if ("defect" in h or "story" in h) and "parent" not in h and ("id" in h or "#" in h): d["id"] = vs
                    elif "issue description" in h: d["desc"] = vs[:120]
                    elif h == "priority": d["priority"] = vs
                    elif h == "status": d["status"] = vs
                    elif "issue type" in h: d["issue_type"] = vs
                if d.get("id") and d["id"] != "nan" and "NO Defect" not in d.get("desc", ""):
                    defects.append(d)
    wb.close()

    # Metrics & Alerts
    dl = (end_dt - today).days
    td = (end_dt - start_dt).days if isinstance(start_dt, datetime) else 14
    pe = min(100, max(0, int((today - start_dt).days / max(td, 1) * 100))) if isinstance(start_dt, datetime) else 50
    total_count = len(stories)
    completed_count = sum(1 for s in stories if s["status"] == "Completed")
    done_pct = int(completed_count / max(total_count, 1) * 100)

    alerts = []
    for s in stories:
        if s["act_sp"] > 0 and s["est_sp"] > 0 and s["act_sp"] > s["est_sp"] * 1.5:
            alerts.append({"sev": "high", "type": "OVERRUN", "msg": f"{s['id']}: Actual ({s['act_sp']}) vs Est ({s['est_sp']})"})
    if pe > 50:
        for s in stories:
            if s["status"] == "Not Started":
                alerts.append({"sev": "high" if pe > 70 else "medium", "type": "BLOCKED",
                               "msg": f"{s['id']}: Not Started at {pe}% elapsed"})
    for s in stories:
        for rf in s.get("review_flags", []):
            alerts.append({"sev": "medium", "type": "REVIEW", "msg": f"{s['id']}: {rf[:140]}"})
    for d in defects:
        sv = "high" if "significant" in d.get("priority", "").lower() else "medium"
        alerts.append({"sev": sv, "type": "DEFECT", "msg": f"{d['id']}: {d.get('desc', '')[:100]} [{d.get('priority', '')}]"})
    if dl <= 2 and done_pct < 50 and total_count > 0:
        alerts.append({"sev": "critical", "type": "AT RISK", "msg": f"Only {completed_count}/{total_count} stories done with {max(0, dl)} day(s) left"})

    counts = {}
    for s in stories:
        counts[s["status"]] = counts.get(s["status"], 0) + 1

    # Use filename-derived name as fallback if dashboard has no app name
    app_name = info.get("app", "") or display_name_from_file(filepath)

    return {
        "app": app_name, "sprint": info.get("sprint", ""),
        "start": start_dt.strftime("%Y-%m-%d") if isinstance(start_dt, datetime) else "",
        "end": end_dt.strftime("%Y-%m-%d"), "days_left": dl, "pct_elapsed": pe,
        "counts": counts, "total": total_count, "done_pct": done_pct,
        "defects": defects, "defect_count": len(defects),
        "alerts": alerts, "stories": stories,
    }, None


# ── Discovery ───────────────────────────────────

def discover_workbooks(root_dir):
    """Recursively find all files ending with 'Sprint Metrics.xlsx' or
    'Sprint_Metrics.xlsx' under root_dir. Skips Office temp files.
    Returns list of (display_name, filepath)."""
    root = Path(root_dir)
    found = []
    for f in sorted(root.rglob("*.xlsx")):
        if f.is_file() and not f.name.startswith("~"):
            if re.search(r'Sprint[\s_]*Metrics\.xlsx$', f.name, re.IGNORECASE):
                display = display_name_from_file(f)
                found.append((display, f))
    return found


# ── Email HTML ──────────────────────────────────

def generate_email_html(data):
    M = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    def fmt(s):
        if not s: return ""
        d = datetime.strptime(s, "%Y-%m-%d"); return f"{M[d.month-1]} {d.day}"
    esc = html_mod.escape
    today = datetime.strptime(data["today"], "%Y-%m-%d")
    L = []
    L.append(f"""<div style="font-family:'Segoe UI',Arial,sans-serif;max-width:720px;margin:0 auto;background:#f4f5f7;padding:20px">
<table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:16px"><tr>
<td style="font-size:22px;font-weight:700;color:#111827">Engineering Delivery Overview</td>
<td style="text-align:right;font-size:12px;color:#9ca3af">Generated: {data['generated']}</td></tr></table>""")

    if data.get("errors"):
        rows = "".join(f'<tr><td style="padding:4px 8px;font-weight:600;color:#991b1b;font-size:12px;border-bottom:1px solid #fecaca;white-space:nowrap">{esc(e["app"])}</td><td style="padding:4px 8px;font-size:12px;color:#92400e;border-bottom:1px solid #fecaca">{esc(e["reason"])}</td></tr>' for e in data["errors"])
        L.append(f"""<div style="background:#fef2f2;border:1px solid #fecaca;border-radius:8px;padding:10px 14px;margin-bottom:14px">
<div style="font-size:12px;font-weight:700;color:#991b1b;margin-bottom:6px">⚠ {len(data["errors"])} Application{"s" if len(data["errors"])>1 else ""} Not Reporting</div>
<table width="100%" cellpadding="0" cellspacing="0">{rows}</table></div>""")

    tt=len(data["teams"]); td=sum(t["defect_count"] for t in data["teams"])
    L.append(f"""<table cellpadding="0" cellspacing="0" style="width:100%;margin-bottom:16px"><tr>
<td style="background:#fff;border:1px solid #e3e5ea;border-radius:10px;padding:14px 20px;width:49%"><div style="font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:1px;color:#9ca3af">Teams</div><div style="font-size:28px;font-weight:700">{tt}</div></td><td style="width:2%"></td>
<td style="background:#fff;border:1px solid #e3e5ea;border-radius:10px;padding:14px 20px;width:49%"><div style="font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:1px;color:#9ca3af">Defects</div><div style="font-size:28px;font-weight:700;color:{'#dc2626' if td else '#111827'}">{td}</div></td></tr></table>""")

    gm={}
    for t in data["teams"]: k=t["start"]+"|"+t["end"]; gm.setdefault(k,[]).append(t)
    gks=sorted(gm.keys(),key=lambda x:x.split("|")[1])
    sc={"Completed":("#059669","#ecfdf5"),"Testing":("#2563eb","#eff6ff"),"Review":("#7c3aed","#f5f3ff"),"In Progress":("#d97706","#fffbeb"),"Not Started":("#6b7280","#f3f4f6")}
    csn={"Completed":"Done","Testing":"Test","Review":"Rev","In Progress":"Dev","Not Started":"NS"}
    co=["Completed","Testing","Review","In Progress","Not Started"]
    at={"DEFECT":"defect","OVERRUN":"overrun","BLOCKED":"blocked","AT RISK":"atrisk","REVIEW":"review"}
    fc={"defect":("#dc2626","#fef2f2"),"overrun":("#d97706","#fffbeb"),"blocked":("#d97706","#fffbeb"),"atrisk":("#dc2626","#fef2f2"),"review":("#a16207","#fefce8")}
    fl={"defect":"Defect","overrun":"Overrun","blocked":"Blocked","atrisk":"At Risk","review":"Review"}

    for gk in gks:
        gs,ge=gk.split("|"); teams=gm[gk]; dlv=(datetime.strptime(ge,"%Y-%m-%d")-today).days
        dl_bg="#fef2f2" if dlv<=0 else "#fffbeb" if dlv<=3 else "#ecfdf5"
        dl_fg="#dc2626" if dlv<=0 else "#d97706" if dlv<=3 else "#059669"
        dl_txt="Ends Today" if dlv<=0 else f"{dlv}d left"
        L.append(f"""<div style="margin-bottom:18px"><table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:6px"><tr>
<td style="font-size:13px;font-weight:600;color:#111827">{fmt(gs)} → {fmt(ge)}</td>
<td style="width:80px"><span style="font-size:10px;font-weight:600;padding:3px 8px;border-radius:10px;background:{dl_bg};color:{dl_fg}">{dl_txt}</span></td>
<td style="text-align:right;font-size:11px;color:#9ca3af">{len(teams)} team{"s" if len(teams)>1 else ""}</td></tr></table>""")
        for t in teams:
            fbt={}
            for a in t["alerts"]: fbt.setdefault(at.get(a["type"],"other"),[]).append(a)
            hi=any(ft in fbt for ft in ["defect","overrun","blocked","atrisk"]); bc="#dc2626" if hi else "#e3e5ea"
            pills="".join(f'<span style="display:inline-block;padding:3px 8px;border-radius:5px;background:{sc[c][1]};color:{sc[c][0]};font-size:10px;font-weight:600;margin-right:2px">{t["counts"].get(c,0)} {csn[c]}</span>' for c in co)
            fh=""
            for ft in ["defect","overrun","blocked","atrisk","review"]:
                items=fbt.get(ft,[])
                if not items: continue
                cl=f'{len(items)} {fl[ft]}{"s" if len(items)>1 and ft!="atrisk" else ""}' if ft!="atrisk" else "At Risk"
                fh+=f'<span style="display:inline-block;padding:2px 8px;border-radius:10px;background:{fc[ft][1]};color:{fc[ft][0]};font-size:10px;font-weight:600;margin-right:3px">● {cl}</span>'
            if not fh: fh='<span style="display:inline-block;padding:2px 8px;border-radius:10px;background:#ecfdf5;color:#059669;font-size:10px;font-weight:600">✓ On Track</span>'
            dw,tw=max(1,t["done_pct"]),max(1,min(100,t["pct_elapsed"]))
            L.append(f"""<div style="background:#fff;border:1px solid #e3e5ea;border-left:4px solid {bc};border-radius:10px;margin-bottom:6px;padding:12px 14px">
<table width="100%" cellpadding="0" cellspacing="0"><tr>
<td style="vertical-align:top;width:150px"><div style="font-size:14px;font-weight:700;color:#111827">{esc(t["app"])}</div><div style="font-size:11px;color:#9ca3af;margin-top:2px">{esc(t["sprint"])}</div></td>
<td style="vertical-align:top;padding:0 12px;width:220px"><table width="100%" cellpadding="0" cellspacing="0">
<tr><td style="font-size:10px;font-weight:600;color:#4b5563;width:34px">Done:</td><td><div style="background:#f4f5f7;border-radius:4px;height:7px;width:100%"><div style="background:#10b981;border-radius:4px;height:7px;width:{dw}%"></div></div></td><td style="font-size:10px;font-weight:600;color:#4b5563;width:34px;text-align:right">{t["done_pct"]}%</td></tr>
<tr><td colspan="3" style="height:3px"></td></tr>
<tr><td style="font-size:10px;font-weight:600;color:#4b5563;width:34px">Time:</td><td><div style="background:#f4f5f7;border-radius:4px;height:7px;width:100%"><div style="background:#94a3b8;border-radius:4px;height:7px;width:{tw}%"></div></div></td><td style="font-size:10px;font-weight:600;color:#4b5563;width:34px;text-align:right">{min(100,t["pct_elapsed"])}%</td></tr>
</table></td><td style="vertical-align:top;text-align:right">{pills}</td></tr></table>
<div style="margin-top:6px">{fh}</div></div>""")
        L.append("</div>")
    L.append("</div>")
    return "\n".join(L)


# ── Outlook Sender ──────────────────────────────

def send_via_outlook(subject, body, to, cc="", attach=None):
    try:
        import win32com.client
    except ImportError:
        print("ERROR: pywin32 not installed. Run:  pip install pywin32")
        print("       Then:  python -m pywin32_postinstall -install"); sys.exit(1)
    try:
        ol = win32com.client.Dispatch("Outlook.Application")
        m = ol.CreateItem(0)
        m.Subject = subject; m.HTMLBody = body; m.To = to
        if cc: m.CC = cc
        if attach and os.path.exists(attach): m.Attachments.Add(os.path.abspath(attach))
        m.Send()
        print(f"  Email sent → {to}")
        if cc: print(f"  CC → {cc}")
    except Exception as e:
        print(f"  ERROR sending: {e}\n  Ensure Outlook desktop is open."); sys.exit(1)


# ── Main ────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Sprint Dashboard Generator")
    parser.add_argument("--config", default=None, help="Path to config.json")
    parser.add_argument("--send", action="store_true", help="Send email via Outlook")
    args = parser.parse_args()

    script_dir = Path(__file__).parent
    cfg = load_config(args.config or str(script_dir / "config.json"))

    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    print(f"Sprint Dashboard Generator — {today.strftime('%B %d, %Y')}")

    wb_dir = Path(cfg["workbooks_dir"])
    out_dir = Path(cfg["output_dir"])
    out_dir.mkdir(parents=True, exist_ok=True)
    smap = cfg["status_map"]
    trivials = set(v.lower() for v in cfg["trivial_review_values"])

    workbooks = discover_workbooks(wb_dir)
    if not workbooks:
        print(f"  No files ending with 'Sprint Metrics.xlsx' found under {wb_dir}"); sys.exit(1)

    print(f"  Root: {wb_dir}")
    print(f"  Found {len(workbooks)} Sprint_Metrics file(s)\n")

    teams, errors = [], []
    for display, fpath in workbooks:
        print(f"  {display}...", end=" ")
        try:
            td, err = extract_team(str(fpath), today, smap, trivials)
            if td:
                teams.append(td); print(f"OK — {td['app']}")
            else:
                errors.append({"app": display, "reason": err}); print(f"ERROR — {err}")
        except Exception as e:
            errors.append({"app": display, "reason": str(e)[:150]}); print(f"EXCEPTION — {e}")

    teams.sort(key=lambda t: t["days_left"])
    data = {"generated": today.strftime("%B %d, %Y"), "today": today.strftime("%Y-%m-%d"),
            "teams": teams, "errors": errors}

    print(f"\n  Active: {len(teams)}  |  Errors: {len(errors)}  |  "
          f"Defects: {sum(t['defect_count'] for t in teams)}  |  "
          f"Alerts: {sum(len(t['alerts']) for t in teams)}")

    tpl_path = script_dir / cfg["template_file"]
    interactive_path = None
    if tpl_path.exists():
        with open(tpl_path, "r", encoding="utf-8") as f: tpl = f.read()
        interactive_path = out_dir / "sprint_dashboard.html"
        with open(interactive_path, "w", encoding="utf-8") as f:
            f.write(tpl.replace("%%DATA%%", json.dumps(data)))
        print(f"\n  Interactive: {interactive_path}")
    else:
        print(f"\n  WARNING: Template not found: {tpl_path}")

    email_html = generate_email_html(data)
    email_path = out_dir / "sprint_dashboard_email.html"
    with open(email_path, "w", encoding="utf-8") as f: f.write(email_html)
    print(f"  Email HTML:  {email_path}")

    if args.send:
        to = cfg["email"]["to"]
        if not to: print("\n  ERROR: email.to is empty in config.json"); sys.exit(1)
        subj = cfg["email"]["subject"].format(date=today.strftime("%B %d, %Y"))
        att = str(interactive_path) if cfg["email"]["attach_interactive"] and interactive_path else None
        print("\n  Sending email...")
        send_via_outlook(subj, email_html, to, cfg["email"]["cc"], att)

    print("\nDone.")


if __name__ == "__main__":
    main()
