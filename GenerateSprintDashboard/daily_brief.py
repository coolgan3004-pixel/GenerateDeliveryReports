#!/usr/bin/env python3
"""
Daily Sprint Status Brief  -- consolidated view across all team workbooks.

For every '<App> - Sprint Metrics.xlsx' under ROOT it:
  1. reads the 'Dashboard' sheet ROW 1 (key/value) for the CURRENT sprint name,
     its Status sheet name, start/end dates and the application name;
  2. opens that sprint's '...-Status' sheet and reads each story
     (Developer | Story ID | Description | Status | Estimated(sp) | Actual(sp) | Dev dates);
  3. derives progress from the Status STAGE text, compares pace vs time elapsed,
     forecasts whether each story / the team will finish on time;
  4. writes one self-contained HTML brief: portfolio summary + per-team cards
     + per-team story drill-down.

Run:  python daily_brief.py
"""
import os, glob, html, math, re
from datetime import date, datetime, timedelta
import openpyxl

# ---------------------------------------------------------------- config
# Resolve folder relative to this script so it works in any session/machine.
# Override with the SPRINT_BRIEF_ROOT environment variable if needed.
ROOT  = os.environ.get("SPRINT_BRIEF_ROOT") or os.path.dirname(os.path.abspath(__file__))
OUT   = os.environ.get("SPRINT_BRIEF_OUT") or os.path.join(ROOT, "Daily_Status_Brief.html")
TODAY = date.today()

# Dashboard row-1 key/value positions are matched by KEY TEXT (robust to shifts)
DASH_KEYS = {
    "sprint": ["current sprint"],          # but NOT "current sprint sheet"
    "sheet":  ["current sprint sheet", "sprint sheet"],
    "start":  ["sprint start"],
    "end":    ["sprint end"],
    "app":    ["application name", "application name "],
}

# status-sheet column synonyms
COL = {
    "dev":    ["developer", "owner", "assignee"],
    "id":     ["story id", "story id/defect", "id/defect", "defect id", "id"],
    "desc":   ["description", "summary", "title"],
    "status": ["status"],
    "est":    ["estimated(sp)", "estimated (sp)", "estimated(hrs/sp)", "estimated", "story points", "estimate"],
    "act":    ["actual(sp)", "actual (sp)", "actual(hrs/sp)", "actual"],
    "dstart": ["dev start date", "dev start"],
    "dend":   ["dev end date", "dev end"],
    "sprintcol": ["sprint"],                 # the per-row Sprint label (may say "... Stretch")
    "main":   ["main sprint", "main/stretch"],
}

# stage text -> (canonical stage label, progress weight 0..1)
def stage_of(status):
    t = re.sub(r"\s+", " ", str(status or "").strip().lower())
    if not t:                                   return ("Not Started", 0.0)
    if any(w in t for w in ("blocked", "on hold", "hold", "impediment", "blocker")):
        return ("Blocked", 0.40)
    if any(w in t for w in ("approved", "completed", "complete", "fixed", "closed",
                            "done", "signed off", "sign off", "deployed", "released")):
        return ("Done", 1.0)
    if "ready for test" in t or "ready to test" in t:   return ("Ready for Test", 0.80)
    if "test" in t or "qa" in t:                        return ("Testing", 0.85)
    if "review" in t or "crc" in t:                     return ("In Review", 0.75)
    if any(w in t for w in ("develop", "in progress", "inprogress", "being fixed", "wip")):
        return ("In Development", 0.45)
    if any(w in t for w in ("refine", "groom", "analysis", "backlog",
                            "not started", "to do", "todo", "ready for dev", "yet to")):
        return ("Refinement", 0.10)
    return ("In Progress", 0.30)

DONE_STAGES = {"Done"}

# ---------------------------------------------------------------- helpers
def norm(s):  return re.sub(r"\s+", " ", str(s).strip().lower()) if s is not None else ""
def naf(v):   # is this a usable value (not #N/A / blank)
    return v is not None and str(v).strip() not in ("", "#N/A", "#n/a", "None", "0")

def to_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y", "%Y-%m-%d %H:%M:%S"):
        try: return datetime.strptime(str(v).strip(), fmt).date()
        except ValueError: continue
    return None

def num(v):
    try:
        f = float(str(v).strip())
        return f if not math.isnan(f) else None
    except (ValueError, TypeError): return None

def working_days(d0, d1):
    if not d0 or not d1 or d1 < d0: return 0
    n, cur = 0, d0
    while cur <= d1:
        if cur.weekday() < 5: n += 1
        cur += timedelta(days=1)
    return n

def add_working_days(start, k):
    cur, added = start, 0
    k = max(0, int(round(k)))
    while added < k:
        cur += timedelta(days=1)
        if cur.weekday() < 5: added += 1
    return cur

def team_name(path):
    b = os.path.basename(path)
    return re.sub(r"\s*-\s*sprint metrics.*$", "", b, flags=re.I).strip() or os.path.splitext(b)[0]

def find_col(header, keys):
    cells = [norm(c) for c in header]
    for k in keys:
        for i, c in enumerate(cells):
            if c == k: return i
    for k in keys:                                  # fuzzy contains
        for i, c in enumerate(cells):
            if k in c or c in k and len(c) > 2: return i
    return None

# ---------------------------------------------------------------- dashboard row1
def parse_dashboard_header(ws):
    r1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    out = {}
    for j in range(0, len(r1) - 1):
        k = norm(r1[j])
        if not k: continue
        v = r1[j + 1]
        # match longest/specific keys first to avoid 'current sprint' eating 'current sprint sheet'
        if k in ("current sprint sheet", "sprint sheet") and "sheet" not in out:
            out["sheet"] = v
        elif k == "current sprint" and "sprint" not in out:
            out["sprint"] = v
        elif "sprint start" in k and "start" not in out:
            out["start"] = v
        elif "sprint end" in k and "end" not in out:
            out["end"] = v
        elif "application name" in k and "app" not in out:
            out["app"] = v
    return out

def locate_status_sheet(wb, sheet_hint, sprint_name):
    names = wb.sheetnames
    if naf(sheet_hint):
        h = str(sheet_hint).strip()
        for sn in names:
            if sn.strip() == h: return sn
        for sn in names:
            if sn.strip().lower() == h.lower(): return sn
        for sn in names:                                  # startswith trimmed
            if sn.strip().lower().startswith(h.lower()[:14]): return sn
    if naf(sprint_name):                                  # fall back: "<sprint>-Status"
        sp = str(sprint_name).strip().lower()
        cands = [sn for sn in names if "status" in sn.lower() and sn.strip().lower().startswith(sp[:10])]
        if cands: return cands[0]
    return None

STAGE_HINTS = ("story ", "defect ", "completed", "complete", "review", "tested",
               "testing", "develop", "refine", "approved", "ready for", "blocked",
               "fixed", "peer", "in progress", "qa", "crc", "not started", "to do")

def detect_status_col(rows, exclude):
    """Find the column whose cell values look like workflow statuses."""
    ncol = max((len(r) for r in rows), default=0)
    best, best_score = None, 0
    for j in range(ncol):
        if j in exclude:
            continue
        score = 0
        for r in rows[1:25]:
            v = norm(r[j]) if j < len(r) else ""
            if v and any(h in v for h in STAGE_HINTS):
                score += 1
        if score > best_score:
            best, best_score = j, score
    return best if best_score >= 2 else None

# ---------------------------------------------------------------- read one workbook
def read_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    rec = {"team": team_name(path), "path": path, "app": "", "sprint": "",
           "start": None, "end": None, "stories": [], "note": "", "ok": False}
    if "Dashboard" not in wb.sheetnames:
        rec["note"] = "No Dashboard sheet."
        return rec
    hdr = parse_dashboard_header(wb["Dashboard"])
    rec["app"]    = str(hdr.get("app") or "").strip()
    rec["sprint"] = str(hdr.get("sprint") or "").strip()
    rec["start"]  = to_date(hdr.get("start"))
    rec["end"]    = to_date(hdr.get("end"))
    if not naf(hdr.get("sprint")):
        rec["note"] = "No current sprint set in Dashboard (#N/A)."
        return rec

    sheet = locate_status_sheet(wb, hdr.get("sheet"), hdr.get("sprint"))
    if not sheet:
        rec["note"] = f"Current-sprint status sheet not found ({hdr.get('sheet')})."
        return rec

    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        rec["note"] = "Status sheet empty."; return rec
    header = rows[0]
    ci = {k: find_col(header, v) for k, v in COL.items()}
    if ci.get("status") is None:
        ex = {ci.get("desc"), ci.get("id")}
        ex.discard(None)
        ci["status"] = detect_status_col(rows, ex)
        if ci.get("status") is None and ci.get("desc") is not None:
            ci["status"] = ci["desc"] + 1
    stories = []
    for r in rows[1:]:
        if not any(naf(c) for c in r): continue
        def g(key):
            i = ci.get(key)
            return r[i] if i is not None and i < len(r) else None
        sid = g("id"); desc = g("desc"); status = g("status")
        if not naf(sid) and not naf(desc):       # skip blank / total rows
            continue
        if isinstance(sid, str) and sid.strip().lower() in ("s.no", "total", "totals"):
            continue
        stage, weight = stage_of(status)
        est = num(g("est"))
        is_stretch = (norm(g("main")) == "stretch") or ("stretch" in norm(g("sprintcol")))
        stories.append({
            "dev": str(g("dev") or "").strip(),
            "id": str(sid or "").strip(),
            "desc": str(desc or "").strip(),
            "status_raw": str(status or "").strip(),
            "stage": stage, "weight": weight,
            "est": est if est else 0.0,
            "dend": to_date(g("dend")),
            "stretch": is_stretch,
        })
    rec["stories"] = stories
    rec["ok"] = True
    if rec["end"] and rec["end"] < TODAY:
        days = (TODAY - rec["end"]).days
        rec["note"] = f"Sprint ended {days} day(s) ago — data may be stale or in close-out."
    elif rec["start"] and rec["start"] > TODAY:
        rec["note"] = "Sprint has not started yet."
    return rec

# ---------------------------------------------------------------- analytics
def classify(rec):
    all_stories = rec["stories"]
    stories = [s for s in all_stories if not s.get("stretch")]   # committed/main only
    stretch = [s for s in all_stories if s.get("stretch")]
    use_pts = any(s["est"] for s in stories)
    def w(s): return s["est"] if use_pts else 1.0
    total = sum(w(s) for s in stories) or 0.0
    done  = sum(w(s) for s in stories if s["stage"] in DONE_STAGES)
    weighted = sum(w(s) * s["weight"] for s in stories)
    prog = (weighted / total) if total else 0.0
    done_frac = (done / total) if total else 0.0

    start, end = rec["start"], rec["end"]
    total_wd = working_days(start, end) or 1
    elapsed_wd = min(working_days(start, TODAY), total_wd) if start else total_wd
    time_frac = elapsed_wd / total_wd if total_wd else 1.0
    days_left = working_days(TODAY, end) if end else 0

    # pace: fraction of plan done per fraction of time -> projected % by sprint end
    if time_frac > 0:
        proj = min(2.0, prog / time_frac)
    else:
        proj = 1.0 if prog >= 1 else 0.0
    proj_pct = round(proj * 100)

    blocked = sum(1 for s in stories if s["stage"] == "Blocked")

    # per-story flags (compute for all rows incl. stretch, for display)
    for s in all_stories:
        s.update(story_flag(s, time_frac, end, days_left))
    off = sum(1 for s in stories if s["flag"] == "Off Track")
    risk = sum(1 for s in stories if s["flag"] == "At Risk")

    overdue = bool(end and end < TODAY and done_frac < 0.999)
    if total == 0:
        verdict = "No Data"
    elif done_frac >= 0.999:
        verdict = "On Track"          # fully done
    elif overdue:
        verdict = "Off Track"
    elif blocked or off >= 2 or proj_pct < 85:
        verdict = "Off Track" if (off >= 2 or proj_pct < 70) else "At Risk"
    elif risk or off or proj_pct < 100:
        verdict = "At Risk"
    else:
        verdict = "On Track"

    rec["roll"] = {
        "use_pts": use_pts, "total": total, "done": done,
        "prog_pct": round(prog * 100), "done_pct": round(done_frac * 100),
        "time_pct": round(time_frac * 100), "proj_pct": proj_pct,
        "days_left": days_left, "blocked": blocked, "off": off, "risk": risk,
        "n": len(stories), "n_done": sum(1 for s in stories if s["stage"] in DONE_STAGES),
        "stretch_n": len(stretch),
        "stretch_done": sum(1 for s in stretch if s["stage"] in DONE_STAGES),
        "verdict": verdict, "overdue": overdue,
    }
    return rec

def story_flag(s, time_frac, end, days_left):
    if s["stage"] in DONE_STAGES:
        return {"flag": "Done", "eta": end, "pct": 100}
    if s["stage"] == "Blocked":
        return {"flag": "Blocked", "eta": None, "pct": round(s["weight"] * 100)}
    weight = s["weight"]
    # forecast finish: prefer recorded Dev End Date when it's still ahead
    eta = s["dend"] if (s["dend"] and s["dend"] >= TODAY) else None
    if eta is None and end:
        remaining = max(0.0, 1 - weight)
        # spread remaining work across remaining sprint time, min 1 day
        eta = add_working_days(TODAY, max(1, remaining * max(days_left, 1)))
    late_by_date = bool(s["dend"] and end and s["dend"] > end)
    expected = time_frac
    if weight >= expected - 0.10 and not late_by_date and (not end or not eta or eta <= end):
        flag = "On Track"
    elif weight >= expected - 0.25 and not late_by_date:
        flag = "At Risk"
    else:
        flag = "Off Track"
    if weight <= 0.10 and time_frac > 0.55:
        flag = "Off Track"
    if late_by_date and flag == "On Track":
        flag = "At Risk"
    return {"flag": flag, "eta": eta, "pct": round(weight * 100)}

# ---------------------------------------------------------------- HTML / output
CLR = {"On Track": "#1a7f37", "At Risk": "#bf8700", "Off Track": "#cf222e",
       "Blocked": "#8250df", "Done": "#57606a", "No Data": "#8c959f"}
BG  = {"On Track": "#e6f4ea", "At Risk": "#fff4d6", "Off Track": "#ffe3e3",
       "Blocked": "#f1e9fc", "Done": "#eef1f4", "No Data": "#eef1f4"}
ORD = {"Off Track": 0, "At Risk": 1, "Blocked": 1, "On Track": 2, "Done": 3, "No Data": 4}

# every selector scoped under .dsb-root so the brief can be injected into any page
STYLE = """
.dsb-root,.dsb-root *{box-sizing:border-box}
.dsb-root{font-family:-apple-system,Segoe UI,Roboto,Helvetica,Arial,sans-serif;color:#1f2328;
 background:#f6f8fa;padding:0;margin:0}
.dsb-root .wrap{max-width:1180px;margin:0 auto;padding:24px 20px 60px}
.dsb-root h1{font-size:24px;margin:0 0 2px}
.dsb-root .sub{color:#656d76;font-size:13px}
.dsb-root .kpis{display:flex;flex-wrap:wrap;gap:12px;margin:18px 0 26px}
.dsb-root .kpi{background:#fff;border:1px solid #d0d7de;border-radius:10px;padding:11px 16px;min-width:118px}
.dsb-root .kpi b{font-size:22px;display:block;line-height:1.1}
.dsb-root .kpi span{font-size:11.5px;color:#656d76}
.dsb-root .cards{display:grid;grid-template-columns:repeat(auto-fill,minmax(270px,1fr));gap:13px;margin-bottom:30px}
.dsb-root .card{background:#fff;border:1px solid #d0d7de;border-left-width:5px;border-radius:10px;padding:13px 15px}
.dsb-root .card h3{margin:0;font-size:15.5px}
.dsb-root .meta{color:#656d76;font-size:11.5px;margin:3px 0 9px}
.dsb-root .m{font-size:11.5px;color:#424a53;margin:3px 0}
.dsb-root h2.sec{font-size:15px;border-bottom:2px solid #d0d7de;padding-bottom:6px;margin:28px 0 14px}
.dsb-root .team{background:#fff;border:1px solid #d0d7de;border-radius:10px;overflow:hidden;margin-bottom:16px}
.dsb-root .th{padding:11px 15px;display:flex;justify-content:space-between;align-items:center;gap:10px;border-bottom:1px solid #eaecef;cursor:pointer}
.dsb-root .th h3{margin:0;font-size:16px}
.dsb-root .th .meta{margin:2px 0 0}
.dsb-root table{border-collapse:collapse;width:100%;font-size:12.5px}
.dsb-root th,.dsb-root td{text-align:left;padding:7px 11px;border-bottom:1px solid #eef0f2;vertical-align:middle}
.dsb-root th{background:#f6f8fa;font-size:10.5px;text-transform:uppercase;letter-spacing:.4px;color:#656d76}
.dsb-root .late{color:#cf222e;font-weight:600}
.dsb-root .note{font-size:11.5px;color:#9a6700;background:#fff8e6;padding:4px 10px;border-top:1px solid #f0e3b8}
.dsb-root details>summary{list-style:none}
.dsb-root details>summary::-webkit-details-marker{display:none}
.dsb-root .inact td{color:#656d76}
.dsb-root{scroll-behavior:smooth}
.dsb-root a.cardlink{text-decoration:none;color:inherit;display:block}
.dsb-root .card{transition:box-shadow .12s,transform .12s}
.dsb-root a.cardlink:hover .card{box-shadow:0 3px 12px rgba(0,0,0,.13);transform:translateY(-1px)}
.dsb-root .missing{background:#fff0f0;border:1px solid #ffc9c9;border-left:4px solid #cf222e;color:#86181d;border-radius:8px;padding:9px 13px;margin:14px 0 2px;font-size:12.5px;line-height:1.6}
.dsb-root .m.stretch{color:#8250df}
.dsb-root .stag{background:#f1e9fc;color:#8250df;font-size:9.5px;font-weight:700;padding:1px 6px;border-radius:8px;text-transform:uppercase;letter-spacing:.3px;margin-left:5px}
.dsb-root tr.strow td{background:#faf8ff}
.dsb-root .team:target{box-shadow:0 0 0 2px #0969da;border-color:#0969da}
"""

def chip(f):
    return ('<span style="background:%s;color:%s;padding:2px 9px;border-radius:11px;'
            'font-size:11.5px;font-weight:600;white-space:nowrap">%s</span>' % (BG[f], CLR[f], f))

def bar(pct, color="#0969da", w=120):
    pct = max(0, min(100, pct))
    return ('<span style="background:#e7ebef;border-radius:6px;height:8px;width:%dpx;display:inline-block;vertical-align:middle">'
            '<span style="background:%s;height:8px;border-radius:6px;width:%d%%;display:inline-block"></span></span>'
            % (w, color, pct))

def esc(x): return html.escape(str(x))

def slug(t): return re.sub(r"[^a-z0-9]+", "-", str(t).lower()).strip("-") or "team"

def split_active(recs):
    active = [r for r in recs if r.get("ok") and r["roll"]["n"]]
    inactive = [r for r in recs if not (r.get("ok") and r["roll"]["n"])]
    active.sort(key=lambda r: (r["end"] or date.max, ORD[r["roll"]["verdict"]]))
    return active, inactive

# ---------------------------------------------------------------- inner HTML
def build_inner(recs):
    active, inactive = split_active(recs)
    tally = {k: 0 for k in ("On Track", "At Risk", "Off Track")}
    for r in active:
        v = r["roll"]["verdict"]
        if v in tally: tally[v] += 1
    tot_stories = sum(r["roll"]["n"] for r in active)
    tot_done = sum(r["roll"]["n_done"] for r in active)
    tot_blocked = sum(r["roll"]["blocked"] for r in active)

    P = ['<div class="wrap">']
    P.append("<h1>Daily Sprint Status Brief</h1>")
    P.append('<div class="sub">%s &middot; %d active team(s), %d inactive</div>'
             % (TODAY.strftime("%A, %d %B %Y"), len(active), len(inactive)))
    if inactive:
        items = []
        for r in inactive:
            reason = r.get("note") or "no current sprint set"
            items.append("<b>%s</b> &ndash; %s" % (esc(r["team"]), esc(reason)))
        P.append('<div class="missing"><b>&#9888; %d team(s) missing a current sprint:</b> %s</div>'
                 % (len(inactive), "; ".join(items)))
    P.append('<div class="kpis">')
    for label, val, col in [("On Track", tally["On Track"], CLR["On Track"]),
                            ("At Risk", tally["At Risk"], CLR["At Risk"]),
                            ("Off Track", tally["Off Track"], CLR["Off Track"]),
                            ("Stories done", "%d/%d" % (tot_done, tot_stories), "#1f2328"),
                            ("Blocked", tot_blocked, CLR["Blocked"])]:
        P.append('<div class="kpi"><b style="color:%s">%s</b><span>%s</span></div>' % (col, val, label))
    P.append('</div>')

    P.append('<div class="cards">')
    for r in active:
        ro = r["roll"]; v = ro["verdict"]
        app_extra = (' &middot; ' + esc(r['app'])) if r['app'] and r['app'].lower() != r['team'].lower() else ''
        sl = slug(r['team'])
        end_txt = r['end'].strftime('%d %b') if r['end'] else 'no date'
        P.append('<a class="cardlink" href="#detail-%s"><div class="card" style="border-left-color:%s">' % (sl, CLR[v]))
        P.append('<div style="display:flex;justify-content:space-between;align-items:center;gap:8px">'
                 '<h3>%s</h3>%s</div>' % (esc(r['team']), chip(v)))
        P.append('<div class="meta">%s%s &middot; ends %s</div>' % (esc(r['sprint']), app_extra, end_txt))
        P.append('<div class="m">Progress %s <b>%d%%</b> &middot; %d/%d stories done</div>'
                 % (bar(ro['prog_pct'], CLR[v]), ro['prog_pct'], ro['n_done'], ro['n']))
        if ro.get('stretch_n'):
            P.append('<div class="m stretch">+%d stretch (%d done) &middot; not counted in progress</div>'
                     % (ro['stretch_n'], ro['stretch_done']))
        P.append('<div class="m">Time elapsed %s %d%%</div>' % (bar(ro['time_pct'], '#8c959f'), ro['time_pct']))
        P.append('<div class="m">Projected at sprint end: <b>%d%%</b> &middot; %d wd left &middot; %d blocked</div>'
                 % (ro['proj_pct'], ro['days_left'], ro['blocked']))
        P.append('</div></a>')
    P.append('</div>')

    P.append('<h2 class="sec">Team detail</h2>')
    for r in active:
        ro = r["roll"]; v = ro["verdict"]
        sub = '%s &middot; %s &rarr; %s' % (esc(r["sprint"]), r["start"], r["end"])
        P.append('<div class="team" id="detail-%s"><details open><summary>' % slug(r['team']))
        P.append('<div class="th"><div><h3>%s</h3><div class="meta">%s</div></div>'
                 '<div>%s &nbsp;<span class="sub">%d%% done / %d%% time</span></div></div></summary>'
                 % (esc(r["team"]), sub, chip(v), ro["prog_pct"], ro["time_pct"]))
        if r["note"]:
            P.append('<div class="note">&#9888; %s</div>' % esc(r["note"]))
        P.append('<table><tr><th>Story</th><th>Description</th><th>Owner</th>'
                 '<th>Stage</th><th>Progress</th><th>Forecast finish</th><th>Status</th></tr>')
        ordered = sorted(r["stories"], key=lambda x: (x.get("stretch", False), ORD.get(x["flag"], 9)))
        for s in ordered:
            eta = s["eta"]
            if s["flag"] == "Done":
                etxt = "&#10003;"
            elif eta is None:
                etxt = "&mdash;"
            else:
                late = bool(r["end"] and eta > r["end"])
                etxt = '<span class="%s">%s%s</span>' % ("late" if late else "",
                        eta.strftime("%d %b"), " (late)" if late else "")
            tag = ' <span class="stag">stretch</span>' if s.get("stretch") else ''
            rowcls = ' class="strow"' if s.get("stretch") else ''
            P.append('<tr%s><td><b>%s</b>%s</td><td>%s</td><td>%s</td>'
                     '<td>%s <span class="sub">(%s)</span></td>'
                     '<td>%s %d%%</td><td>%s</td><td>%s</td></tr>'
                     % (rowcls, esc(s["id"]) or "&ndash;", tag, esc(s["desc"][:70]), esc(s["dev"]),
                        esc(s["stage"]), esc(s["status_raw"][:22]),
                        bar(s["pct"], CLR[s["flag"]], 80), s["pct"], etxt, chip(s["flag"])))
        P.append('</table></details></div>')

    P.append('<div class="sub" style="margin-top:24px">Generated %s '
             '&middot; progress inferred from each story\'s workflow stage; forecast uses recorded Dev End Date '
             'where present, else remaining-stage &times; time left.</div>' % datetime.now().strftime("%Y-%m-%d %H:%M"))
    P.append('</div>')
    return "".join(P)

def render(recs, mode="standalone"):
    inner = build_inner(recs)
    if mode == "fragment":
        return '<style>%s</style>\n<div class="dsb-root">%s</div>' % (STYLE, inner)
    return ('<!doctype html><html><head><meta charset="utf-8">'
            '<meta name="viewport" content="width=device-width,initial-scale=1">'
            '<title>Daily Sprint Status Brief</title><style>%s</style></head>'
            '<body><div class="dsb-root">%s</div></body></html>' % (STYLE, inner))

# ---------------------------------------------------------------- structured data + flags
def build_data(recs):
    active, inactive = split_active(recs)
    def isod(d): return d.isoformat() if d else None
    teams = []
    for r in active:
        ro = r["roll"]
        late = []
        for s in r["stories"]:
            if s.get("stretch"):
                continue
            if s["flag"] in ("Off Track", "At Risk", "Blocked"):
                late.append({"id": s["id"], "description": s["desc"], "owner": s["dev"],
                             "stage": s["stage"], "status": s["status_raw"], "flag": s["flag"],
                             "forecast_finish": isod(s["eta"]),
                             "late": bool(r["end"] and s["eta"] and s["eta"] > r["end"])})
        flagged = ro["verdict"] in ("Off Track", "At Risk") or ro["blocked"] > 0
        teams.append({
            "team": r["team"], "application": r["app"], "sprint": r["sprint"],
            "start": isod(r["start"]), "end": isod(r["end"]),
            "verdict": ro["verdict"], "progress_pct": ro["prog_pct"], "time_elapsed_pct": ro["time_pct"],
            "projected_pct": ro["proj_pct"], "working_days_left": ro["days_left"],
            "stories_total": ro["n"], "stories_done": ro["n_done"], "blocked": ro["blocked"],
            "stretch_total": ro.get("stretch_n", 0), "stretch_done": ro.get("stretch_done", 0),
            "overdue": ro["overdue"], "note": r["note"], "flagged": flagged,
            "attention_stories": late,
        })
    tally = {k: sum(1 for t in teams if t["verdict"] == k) for k in ("On Track", "At Risk", "Off Track")}
    return {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "date": TODAY.isoformat(),
        "summary": {
            "active_teams": len(teams), "inactive_teams": len(inactive),
            "on_track": tally["On Track"], "at_risk": tally["At Risk"], "off_track": tally["Off Track"],
            "stories_total": sum(t["stories_total"] for t in teams),
            "stories_done": sum(t["stories_done"] for t in teams),
            "blocked": sum(t["blocked"] for t in teams),
            "flagged_teams": sum(1 for t in teams if t["flagged"]),
        },
        "teams": teams,
        "inactive": [{"team": r["team"], "sprint": r["sprint"], "note": r["note"]} for r in inactive],
    }

def flag_report(data):
    """Plain-text delay summary (Off Track / At Risk / blocked)."""
    flagged = [t for t in data["teams"] if t["flagged"]]
    flagged.sort(key=lambda t: ORD.get(t["verdict"], 9))
    lines = []
    s = data["summary"]
    lines.append("Daily Sprint Brief — %s" % data["date"])
    lines.append("%d active teams: %d on track, %d at risk, %d off track, %d blocked story-owners."
                 % (s["active_teams"], s["on_track"], s["at_risk"], s["off_track"], s["blocked"]))
    if not flagged:
        lines.append("No delays flagged — all teams on track.")
        return "\n".join(lines)
    lines.append("")
    lines.append("FLAGGED (%d):" % len(flagged))
    for t in flagged:
        bits = "%d%% done vs %d%% time, proj %d%%, %d wd left" % (
            t["progress_pct"], t["time_elapsed_pct"], t["projected_pct"], t["working_days_left"])
        extra = []
        if t["blocked"]: extra.append("%d blocked" % t["blocked"])
        if t["overdue"]: extra.append("sprint ended/overdue")
        tag = (" [" + ", ".join(extra) + "]") if extra else ""
        lines.append("  - %s — %s (%s) — %s%s" % (t["team"], t["verdict"], t["sprint"], bits, tag))
        for st in t["attention_stories"][:4]:
            fin = st["forecast_finish"] or "n/a"
            lines.append("      * %s %s — %s%s (forecast %s)"
                         % (st["id"], st["owner"], st["flag"],
                            " LATE" if st["late"] else "", fin))
    return "\n".join(lines)

# ---------------------------------------------------------------- main
def discover():
    files = []
    for p in glob.glob(os.path.join(ROOT, "**", "*.xlsx"), recursive=True):
        b = os.path.basename(p)
        if b.startswith("~$"):
            continue
        if "archive" in p.lower():
            continue
        if b.lower().endswith("sprint metrics.xlsx"):
            files.append(p)
    return sorted(files)

def main():
    import json
    files = discover()
    recs = []
    for f in files:
        try:
            r = read_workbook(f)
            if r.get("ok"):
                classify(r)
            recs.append(r)
        except Exception as e:
            recs.append({"team": team_name(f), "path": f, "sprint": "", "app": "",
                         "ok": False, "note": "Read error: %s" % e, "roll": None,
                         "start": None, "end": None, "stories": []})
    base = os.path.splitext(OUT)[0]
    with open(base + ".html", "w", encoding="utf-8") as fh:
        fh.write(render(recs, "standalone"))
    with open(base + ".embed.html", "w", encoding="utf-8") as fh:
        fh.write(render(recs, "fragment"))
    data = build_data(recs)
    with open(base + ".json", "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2)
    print(flag_report(data))
    print("\nWrote: %s.html  /  %s.embed.html  /  %s.json" % (base, base, base))

if __name__ == "__main__":
    main()
